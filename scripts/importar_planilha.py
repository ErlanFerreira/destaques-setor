"""Importa setores, lideres e funcionarios de uma planilha .xlsx para o banco.

Uso:
    python scripts/importar_planilha.py caminho/para/planilha.xlsx

A planilha deve ter uma linha de cabecalho com as colunas (em qualquer ordem):
    setor | lider | cpf | funcionario

Cada linha representa um funcionario, associado ao seu lider, setor e ao CPF
do lider (usado para o lider se identificar no formulario).
Rodar mais de uma vez com a mesma planilha e seguro (nao duplica registros).
"""
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

from app.db import SessionLocal
from app.models import Funcionario, Lider, Setor
from app.seed import init_db


def normalizar(valor):
    return (str(valor).strip() if valor is not None else "")


def normalizar_cpf(valor):
    return re.sub(r"\D", "", str(valor)) if valor is not None else ""


def importar(caminho_planilha: str):
    init_db()
    wb = openpyxl.load_workbook(caminho_planilha, data_only=True)
    ws = wb.active

    cabecalho = [normalizar(c).lower() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    try:
        idx_setor = cabecalho.index("setor")
        idx_lider = cabecalho.index("lider")
        idx_cpf = cabecalho.index("cpf")
        idx_funcionario = cabecalho.index("funcionario")
    except ValueError:
        raise SystemExit("A planilha precisa ter as colunas: setor, lider, cpf, funcionario")

    db = SessionLocal()
    setores_cache = {}
    lideres_cache = {}
    funcionarios_cache = {}
    n_setores = n_lideres = n_funcionarios = 0

    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            nome_setor = normalizar(row[idx_setor])
            nome_lider = normalizar(row[idx_lider])
            cpf_lider = normalizar_cpf(row[idx_cpf])
            nome_funcionario = normalizar(row[idx_funcionario])
            if not nome_setor or not nome_lider or not cpf_lider or not nome_funcionario:
                continue

            setor = setores_cache.get(nome_setor)
            if not setor:
                setor = db.query(Setor).filter(Setor.nome == nome_setor).first()
                if not setor:
                    setor = Setor(nome=nome_setor)
                    db.add(setor)
                    db.flush()
                    n_setores += 1
                setores_cache[nome_setor] = setor

            chave_lider = (nome_setor, nome_lider)
            lider = lideres_cache.get(chave_lider)
            if not lider:
                lider = (
                    db.query(Lider)
                    .filter(Lider.nome == nome_lider, Lider.setor_id == setor.id)
                    .first()
                )
                if not lider:
                    lider = Lider(nome=nome_lider, setor_id=setor.id, cpf=cpf_lider)
                    db.add(lider)
                    db.flush()
                    n_lideres += 1
                elif lider.cpf != cpf_lider:
                    lider.cpf = cpf_lider
                lideres_cache[chave_lider] = lider

            chave_funcionario = (chave_lider, nome_funcionario)
            if chave_funcionario not in funcionarios_cache:
                existente = (
                    db.query(Funcionario)
                    .filter(Funcionario.nome == nome_funcionario, Funcionario.lider_id == lider.id)
                    .first()
                )
                if not existente:
                    db.add(Funcionario(nome=nome_funcionario, setor_id=setor.id, lider_id=lider.id))
                    n_funcionarios += 1
                funcionarios_cache[chave_funcionario] = True

        db.commit()
    finally:
        db.close()

    print(f"Importacao concluida: {n_setores} setor(es), {n_lideres} lider(es), {n_funcionarios} funcionario(s) novo(s).")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        raise SystemExit("Uso: python scripts/importar_planilha.py caminho/para/planilha.xlsx")
    importar(sys.argv[1])
